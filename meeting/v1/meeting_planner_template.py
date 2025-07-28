# Minimal meeting planner Excel workbook generator
# Adds features for UNHCR operations in Poland

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
from icalendar import Calendar, Event
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo


# File paths
xlsx_path = "meeting_planner_template.xlsx"
dir_csv = "directory_template.csv"
rost_csv = "roster_template.csv"
ics_path = "meeting_invite.ics"
log_csv = "decision_log.csv"


def main() -> None:
    """Generate meeting planning templates and a sample workbook."""

    # ---- Build Excel workbook ----
    wb = Workbook()
    
    # ========== Sheet: Settings ==========
    ws = wb.active
    ws.title = "Settings"
    
    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    
    ws["A2"] = "Hours_per_year (editable)"
    ws["B2"] = 1760
    
    ws["A3"] = "Interpreter_hourly_USD (edit)"
    ws["B3"] = 40
    
    ws["A4"] = "USD_to_PLN (edit)"
    ws["B4"] = 4.0
    
    # Grade table
    ws["A6"] = "Grade"
    ws["B6"] = "Annual_USD (editable)"
    ws["C6"] = "Hourly_USD (auto)"
    for cell in ws["6:6"]:
        cell.font = Font(bold=True)
    
    grades = [
    ("P5", 220000),
    ("G7", 77000),
    ("P4", ""),
    ("P3", ""),
    ("G6", ""),
    ("G5", ""),
    ]
    start_row = 7
    for i, (grade, annual) in enumerate(grades, start=start_row):
        ws[f"A{i}"] = grade
        ws[f"B{i}"] = annual if annual != "" else None
        ws[f"C{i}"] = f'=IFERROR(B{i}/$B$2,"")'
    
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    
    ws["A14"] = "Notes"
    ws["A15"] = "• Edit Hours_per_year and Annual_USD. Hourly calculates automatically."
    ws["A16"] = "• Set Interpreter_hourly_USD and USD_to_PLN as needed."
    
    # ========== Sheet: Meeting ==========
    ws2 = wb.create_sheet("Meeting")
    
    ws2["A1"] = "Meeting Planner (Lite)"
    ws2["A1"].font = Font(bold=True, size=14)
    
    labels = [
    ("A3", "Title:"),
    ("A4", "Type (Decision/Coordination/Info):"),
    ("A5", "Decision needed (Yes/No):"),
    ("A6", "Pre-read attached (Yes/No):"),
    ("A7", "Duration (minutes):"),
    ("A8", "Location (city):"),
    ("A9", "Interpreter needed (Yes/No):"),
    ("A10", "Meeting date (YYYY-MM-DD):"),
    ("A11", "Start time (HH:MM):"),
    ]
    for cell, text in labels:
        ws2[cell] = text
        ws2[cell].font = Font(bold=True)
    
    ws2["B4"] = "Decision"
    ws2["B5"] = "Yes"
    ws2["B6"] = "Yes"
    ws2["B7"] = 25
    ws2["B8"] = "Warsaw"
    ws2["B9"] = "No"
    ws2["B10"] = "2025-01-01"
    ws2["B11"] = "09:00"
    
    ws2["E3"] = "Summary"
    ws2["E3"].font = Font(bold=True)
    ws2["E4"] = "Total attendees:"
    ws2["E5"] = "Total cost (USD):"
    ws2["E6"] = "Total cost (PLN):"
    ws2["E7"] = "Necessity score (0–100):"
    ws2["E8"] = "Suggestion:"
    for c in ["E4","E5","E6","E7","E8"]:
        ws2[c].font = Font(bold=True)
        
    ws2["A13"] = "Attendees (up to 10 rows)"
    ws2["A13"].font = Font(bold=True)
    ws2["A15"] = "Name"
    ws2["B15"] = "Email"
    ws2["C15"] = "Grade"
    ws2["D15"] = "Role (A/D/R/C)"
    ws2["E15"] = "Hourly (USD)"
    ws2["F15"] = "Cost for this meeting (USD)"
    for cell in ws2["15:15"]:
        cell.font = Font(bold=True)
    
    for r in range(16, 26):
        ws2[f"E{r}"] = f'=IFERROR(VLOOKUP(C{r},Settings!$A$7:$C$20,3,FALSE),"")'
        ws2[f"F{r}"] = f'=IFERROR(E{r}*$B$7/60,"")'
    
    ws2["E26"] = "Totals:"
    ws2["E26"].font = Font(bold=True)
    ws2["F26"] = '=COUNTA(A16:A25)'
    ws2["F27"] = '=IFERROR(SUM(F16:F25),0)'
    ws2["F28"] = '=IFERROR(F27 + IF($B$9="Yes",Settings!$B$3*$B$7/60,0),0)'
    ws2["F29"] = '=ROUND(F28*Settings!$B$4,2)'
    ws2["E27"] = "Total cost (USD):"
    ws2["E28"] = "With interpreter (USD):"
    ws2["E29"] = "Total cost (PLN):"
    for c in ["E27","E28","E29"]:
        ws2[c].font = Font(bold=True)
    
    ws2["E31"] = "Necessity score:"
    ws2["E31"].font = Font(bold=True)
    ws2["F31"] = '=MAX(0, MIN(100, IF(B5="Yes",50,20) + IF(B6="Yes",15,0) - MAX(0,(F26-5)*5) - IF(B7>50,10,0) ))'
    ws2["E32"] = "Suggestion:"
    ws2["E32"].font = Font(bold=True)
    ws2["F32"] = '=IF(F31<40,"Do async memo", IF(F31<70,"15–30 min huddle (cap 5)","Proceed; cap attendees (≤6)"))'
    
    for col, width in zip(["A","B","C","D","E","F"], [24,32,10,16,18,24]):
        ws2.column_dimensions[col].width = width
    
    thin = Side(border_style="thin", color="DDDDDD")
    for r in range(15, 27):
        for c in range(1, 7):
            ws2.cell(row=r, column=c).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    ws2["A32"] = "How to use (quick):"
    ws2["A32"].font = Font(bold=True)
    ws2["A33"] = "1) In Settings sheet, set Hours_per_year, Annual_USD, interpreter rate and exchange rate."
    ws2["A34"] = "2) On Meeting sheet, fill title, yes/no fields, duration, location, date/time and attendees."
    ws2["A35"] = "3) Interpreter cost is added if needed and PLN totals are calculated."
    ws2["A36"] = "4) Use the suggestion to decide async vs. meeting."
    
    # ========== Sheet: Exports ==========
    ws3 = wb.create_sheet("Exports")
    ws3["A1"] = "Copy the rows below into invites or reports as needed."
    ws3["A1"].font = Font(bold=True)
    ws3["A3"] = "Summary line"
    ws3["B3"] = '=CONCAT("Title: ", IFERROR(Meeting!B3,""), " | Type: ", IFERROR(Meeting!B4,""), " | Duration: ", IFERROR(Meeting!B7,""), " min | Location: ", IFERROR(Meeting!B8,""), " | Attendees: ", IFERROR(Meeting!F26,0), " | Cost: $", TEXT(IFERROR(Meeting!F28,0),"#,##0"), " (", TEXT(IFERROR(Meeting!F29,0),"#,##0"), " PLN)")'
    ws3["A5"] = "Suggestion line"
    ws3["B5"] = '=CONCAT("Suggestion: ", IFERROR(Meeting!F32,""), " | Score: ", IFERROR(Meeting!F31,0))'
    
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 120
    
    wb.save(xlsx_path)
    
    # ---- Create ICS invite ----
    
    cal = Calendar()
    event = Event()
    event.add('summary', ws2['B3'].value or 'Meeting')
    
    dt_str = f"{ws2['B10'].value}T{ws2['B11'].value}"
    start_dt = datetime.fromisoformat(dt_str).replace(tzinfo=ZoneInfo('Europe/Warsaw'))
    
    event.add('dtstart', start_dt)
    
    event.add('dtend', start_dt + timedelta(minutes=int(ws2['B7'].value)))
    
    event.add('location', ws2['B8'].value)
    for r in range(16, 26):
        email = ws2[f'B{r}'].value
        if email:
            event.add('attendee', email)
    cal.add_component(event)
    with open(ics_path, 'wb') as f:
        f.write(cal.to_ical())
    
    # ---- Create CSV templates ----
    
    pd.DataFrame({
    "name": ["", ""],
    "email": ["", ""],
    "unit": ["", ""],
    "location": ["", ""],
    "grade_band": ["P5", "G7"],
    "role_tags": ["", ""],
    "manager_email": ["", ""]
    }).to_csv(dir_csv, index=False)
    
    pd.DataFrame({
    "email": ["", ""],
    "workdays": ["Mon-Fri", "Mon-Fri"],
    "start_local": ["09:00", "09:00"],
    "end_local": ["17:00", "17:00"],
    "timezone": ["Europe/Warsaw", "Europe/Warsaw"],
    "exceptions": ["", ""],
    "travel_windows": ["", ""]
    }).to_csv(rost_csv, index=False)
    
    pd.DataFrame({
    "date": [""],
    "title": [""],
    "decision": [""],
    "owner": [""],
    "attendees": [""],
    "cost_usd": [""],
    "score": [""],
    "suggestion": [""]
    }).to_csv(log_csv, index=False)
    
    print("Templates created:", xlsx_path, dir_csv, rost_csv, ics_path, log_csv)
    
    
if __name__ == "__main__":
    main()
